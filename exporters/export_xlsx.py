# exporters/export_xlsx.py
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def export_items_to_xlsx(items, path):
    wb = Workbook()
    ws = wb.active
    header = ["ID","Title","Username","Password","URL","Tags","Notes","Created At","Modified At"]
    ws.append(header)
    for it in items:
        ws.append([
            it.get("id"),
            it.get("title"),
            it.get("username"),
            it.get("secret"),
            it.get("url"),
            it.get("tags"),
            it.get("notes"),
            it.get("created_at"),
            it.get("modified_at")
        ])
    # autosize
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    wb.save(path)
